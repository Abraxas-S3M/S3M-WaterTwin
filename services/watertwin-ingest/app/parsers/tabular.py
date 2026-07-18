"""Shared, template-driven CSV/XLSX reading for the spreadsheet importer.

This module is the single reading engine behind every templated parser
(equipment, tag mapping, lab methods). It deliberately does the boring, safety-
critical plumbing once so the per-domain parsers only declare a *column contract*:

* **Spreadsheets are read defensively.** ``.xlsx`` is opened with openpyxl in
  ``read_only`` + ``data_only`` mode, so macros are never executed and only cached
  cell *values* (never formulas) are read. Macro-enabled ``.xlsm`` workbooks are
  rejected outright before a single byte is parsed.
* **Encoding is detected with an explicit fallback.** CSV bytes are sniffed
  (BOM, then :mod:`charset_normalizer`); if detection is low-confidence the reader
  falls back to a documented default and emits a warning that the encoding was
  guessed -- it never silently mangles text.
* **Headers are normalized** (case, whitespace, punctuation) and matched against a
  per-column synonym list, so ``"Asset ID"``, ``asset_id`` and ``Asset-Id`` all
  resolve to the same canonical column.
* **Row-level problems are collected, never fatal.** A bad cell flags its own row
  and never discards the rest of the sheet; row numbers are reported 1-based
  (matching the spreadsheet, header = row 1).
* **Formula-injection escaping** (:func:`escape_formula`) is implemented here so
  every export path can reuse the one audited helper.

Nothing in this module writes to any control system; it only reads uploaded
files and produces an in-memory, reviewable :class:`ParseReport`.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from collections.abc import Sequence
from dataclasses import dataclass, field
from typing import Any

from watertwin_engineering import specification_range

__all__ = [
    "IngestError",
    "ColumnSpec",
    "RowIssue",
    "ParseReport",
    "escape_formula",
    "normalize_header",
    "read_rows",
    "parse_table",
    "render_template_csv",
    "DEFAULT_CSV_ENCODING",
    "FORMULA_TRIGGER_CHARS",
]

#: Documented fallback encoding used when detection is inconclusive. Spreadsheet
#: tools on Windows most commonly export CSV as Windows-1252 (a superset of
#: Latin-1), so it is the safest lossless-decoding fallback.
DEFAULT_CSV_ENCODING = "cp1252"

#: Leading characters that make a spreadsheet treat a cell as a formula. Per OWASP
#: CSV-injection guidance we also treat a leading TAB and CR as dangerous.
FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


class IngestError(ValueError):
    """Raised for a fatal, whole-file problem (unreadable / disallowed file)."""


# ---------------------------------------------------------------------------
# Formula-injection escaping (reused by every export path)
# ---------------------------------------------------------------------------


def escape_formula(value: Any) -> Any:
    """Neutralize spreadsheet formula injection for a single exported cell.

    If ``value`` is a string that a spreadsheet could interpret as a formula
    (it begins with one of :data:`FORMULA_TRIGGER_CHARS` -- ``=``, ``+``, ``-``,
    ``@``, TAB or CR, optionally behind leading whitespace) the returned string is
    prefixed with a single quote so the cell is rendered as literal text. Any other
    value (including non-strings and empty strings) is returned unchanged.

    This is the single audited escape helper: every path that *exports* imported
    data back into a CSV/XLSX must route user-supplied cells through it.
    """
    if not isinstance(value, str) or value == "":
        return value
    stripped = value.lstrip()
    if stripped and stripped[0] in FORMULA_TRIGGER_CHARS:
        return "'" + value
    return value


# ---------------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------------


def normalize_header(raw: str) -> str:
    """Normalize a raw header cell to a comparable canonical token.

    Lower-cases, trims, strips a parenthesised unit annotation (``(m3/h)``),
    collapses runs of whitespace/hyphens/dots into single underscores and drops
    remaining punctuation. ``"Rated Flow (m3/h)"`` -> ``"rated_flow"``.
    """
    text = str(raw).strip().lower()
    text = re.sub(r"\(.*?\)", " ", text)  # drop "(m3/h)" style unit annotations
    text = text.replace("%", " pct ")
    text = re.sub(r"[\s\-./\\]+", "_", text)
    text = re.sub(r"[^0-9a-z_]", "", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


# ---------------------------------------------------------------------------
# Column contract
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ColumnSpec:
    """One column in a template contract.

    Attributes:
        name: Canonical column name; also the header written into the template.
        label: Human-readable label for UIs.
        required: Whether a row missing this value is rejected.
        kind: ``"string"``, ``"number"`` or ``"date"``.
        aliases: Accepted header synonyms (matched after normalization).
        unit: Documented default unit for a numeric column (embedded in the
            template header). ``None`` for non-numeric columns.
        range_key: Key into ``watertwin_engineering.SPECIFICATION_RANGES`` used to
            range-check a numeric value. ``None`` skips range validation.
        default: Value substituted when an optional numeric cell is blank.
        example: Example value written into the generated template's sample row.
        description: Documentation for the published contract.
    """

    name: str
    label: str
    required: bool
    kind: str
    aliases: tuple[str, ...] = ()
    unit: str | None = None
    range_key: str | None = None
    default: float | None = None
    example: str = ""
    description: str = ""

    @property
    def header(self) -> str:
        """Header written into the generated template (unit-annotated numbers)."""
        if self.kind == "number" and self.unit:
            return f"{self.name} ({self.unit})"
        return self.name

    def to_contract(self) -> dict[str, Any]:
        """Return the published, serialisable contract entry for this column."""
        entry: dict[str, Any] = {
            "name": self.name,
            "label": self.label,
            "required": self.required,
            "kind": self.kind,
            "description": self.description,
        }
        if self.unit:
            entry["unit"] = self.unit
        if self.range_key:
            entry["range"] = specification_range(self.range_key).describe()
        return entry


# ---------------------------------------------------------------------------
# Result containers
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class RowIssue:
    """One problem found while parsing, tied to a 1-based spreadsheet row."""

    row: int
    severity: str  # "error" | "warning"
    message: str
    column: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "severity": self.severity,
            "message": self.message,
            "column": self.column,
        }


@dataclass
class ParseReport:
    """The reviewable outcome of parsing one uploaded template.

    ``records`` are the fully-validated rows (each stamped with ``provenance``)
    that are safe to import; ``issues`` collects every error and warning with its
    row number so the whole result can be rendered as a review diff. A value that
    is out of range or ambiguous is *never* placed in ``records`` -- it appears in
    ``issues`` instead.
    """

    kind: str
    provenance: str
    columns: list[dict[str, Any]] = field(default_factory=list)
    records: list[dict[str, Any]] = field(default_factory=list)
    issues: list[RowIssue] = field(default_factory=list)

    @property
    def errors(self) -> list[RowIssue]:
        return [i for i in self.issues if i.severity == "error"]

    @property
    def warnings(self) -> list[RowIssue]:
        return [i for i in self.issues if i.severity == "warning"]

    @property
    def ok(self) -> bool:
        """True when nothing errored (warnings are allowed)."""
        return not self.errors

    def add_error(self, row: int, message: str, column: str | None = None) -> None:
        self.issues.append(RowIssue(row=row, severity="error", message=message, column=column))

    def add_warning(self, row: int, message: str, column: str | None = None) -> None:
        self.issues.append(RowIssue(row=row, severity="warning", message=message, column=column))

    def to_dict(self) -> dict[str, Any]:
        return {
            "kind": self.kind,
            "provenance": self.provenance,
            "columns": self.columns,
            "records": self.records,
            "issues": [i.to_dict() for i in self.issues],
            "summary": {
                "records": len(self.records),
                "errors": len(self.errors),
                "warnings": len(self.warnings),
                "ok": self.ok,
            },
        }


@dataclass
class RawTable:
    """Raw rows read from a file before any column mapping/validation."""

    headers: list[str]
    rows: list[list[str]]  # data rows only (header excluded)
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# File reading
# ---------------------------------------------------------------------------


def _extension(filename: str) -> str:
    dot = filename.rfind(".")
    return filename[dot:].lower() if dot >= 0 else ""


def _decode_csv(data: bytes, encoding: str | None) -> tuple[str, list[str]]:
    """Decode CSV bytes to text, returning ``(text, warnings)``.

    Order of preference: an explicit ``encoding`` argument, then a byte-order
    mark, then :mod:`charset_normalizer` detection, then the documented
    :data:`DEFAULT_CSV_ENCODING` fallback (with a warning that it was guessed).
    """
    warnings: list[str] = []
    if encoding:
        return data.decode(encoding, errors="replace"), warnings

    if data.startswith(b"\xef\xbb\xbf"):
        return data.decode("utf-8-sig"), warnings
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        return data.decode("utf-16"), warnings

    # Strict UTF-8 succeeds for the common case without any dependency.
    try:
        return data.decode("utf-8"), warnings
    except UnicodeDecodeError:
        pass

    detected = _detect_encoding(data)
    if detected:
        try:
            text = data.decode(detected)
            warnings.append(
                f"character encoding not declared; guessed {detected!r} via detection"
            )
            return text, warnings
        except (UnicodeDecodeError, LookupError):
            pass

    warnings.append(
        f"character encoding could not be detected; fell back to {DEFAULT_CSV_ENCODING!r} "
        "(non-ASCII text may be wrong -- re-save the file as UTF-8)"
    )
    return data.decode(DEFAULT_CSV_ENCODING, errors="replace"), warnings


def _detect_encoding(data: bytes) -> str | None:
    """Best-effort encoding detection via charset_normalizer (optional)."""
    try:
        from charset_normalizer import from_bytes
    except ImportError:  # pragma: no cover - dependency is declared but be safe
        return None
    match = from_bytes(data).best()
    if match is None:
        return None
    return match.encoding


def _read_csv(data: bytes, encoding: str | None) -> RawTable:
    text, warnings = _decode_csv(data, encoding)
    reader = csv.reader(io.StringIO(text, newline=""))
    all_rows = [list(row) for row in reader]
    if not all_rows:
        return RawTable(headers=[], rows=[], warnings=warnings)
    headers = [c.strip() for c in all_rows[0]]
    return RawTable(headers=headers, rows=all_rows[1:], warnings=warnings)


def _read_xlsx(data: bytes) -> RawTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise IngestError("openpyxl is required to read .xlsx files") from exc

    try:
        # read_only avoids loading the whole workbook into memory; data_only
        # returns cached values instead of formula strings; neither ever
        # evaluates a macro (openpyxl cannot execute VBA at all).
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # openpyxl raises a variety of types on bad input
        raise IngestError(f"could not read spreadsheet: {exc}") from exc

    try:
        sheet = workbook.active
        raw: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            raw.append(["" if cell is None else _stringify_cell(cell) for cell in row])
    finally:
        workbook.close()

    # Trim trailing fully-empty rows produced by spreadsheet padding.
    while raw and all(cell == "" for cell in raw[-1]):
        raw.pop()
    if not raw:
        return RawTable(headers=[], rows=[])
    headers = [c.strip() for c in raw[0]]
    return RawTable(headers=headers, rows=raw[1:])


def _stringify_cell(cell: Any) -> str:
    """Render an openpyxl cell value as the text a parser expects."""
    if isinstance(cell, bool):
        return "true" if cell else "false"
    if isinstance(cell, (_dt.datetime, _dt.date)):
        return cell.isoformat()[:10] if not isinstance(cell, _dt.datetime) else cell.isoformat()
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def read_rows(data: bytes, filename: str, *, encoding: str | None = None) -> RawTable:
    """Read a CSV or XLSX file into a :class:`RawTable`.

    Args:
        data: The raw uploaded file bytes.
        filename: The original filename (used only to pick a reader by extension).
        encoding: Optional explicit CSV encoding; skips detection when provided.

    Raises:
        IngestError: If the file is a macro-enabled ``.xlsm`` (rejected outright),
            an unsupported extension, or is structurally unreadable.
    """
    ext = _extension(filename)
    if ext == ".xlsm":
        raise IngestError(
            "macro-enabled workbooks (.xlsm) are not accepted; re-save as .xlsx or .csv"
        )
    if ext in (".xlsx",):
        return _read_xlsx(data)
    if ext in (".csv", ".txt", ".tsv", ""):
        return _read_csv(data, encoding)
    raise IngestError(f"unsupported file type {ext!r}; upload a .csv or .xlsx template")


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

_NUMBER_RE = re.compile(r"^[-+]?(\d{1,3}(,\d{3})+|\d*)(\.\d+)?([eE][-+]?\d+)?$")
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


@dataclass
class _CellResult:
    value: Any
    error: str | None = None
    warning: str | None = None


def _coerce_number(text: str, spec: ColumnSpec) -> _CellResult:
    """Parse a numeric cell, flagging unit ambiguity as a (non-fatal) warning."""
    cleaned = text.strip()
    if not _NUMBER_RE.match(cleaned):
        # A value carrying a unit token ("50 psi", "1,2e3 bar") is ambiguous: we
        # will not guess -- warn and leave it unparsed per the unit-handling rule.
        return _CellResult(
            value=None,
            warning=(
                f"{spec.name}: could not read {text!r} as a number in "
                f"{spec.unit or 'the documented unit'}; left unparsed (do not embed units)"
            ),
        )
    try:
        value = float(cleaned.replace(",", ""))
    except ValueError:
        return _CellResult(value=None, warning=f"{spec.name}: {text!r} is not a valid number")
    if spec.range_key is not None:
        err = specification_range(spec.range_key).error_for(value)
        if err is not None:
            return _CellResult(value=None, error=err)
    return _CellResult(value=value)


def _coerce_date(text: str, spec: ColumnSpec) -> _CellResult:
    cleaned = text.strip()
    if not _DATE_RE.match(cleaned):
        return _CellResult(
            value=None,
            warning=f"{spec.name}: {text!r} is not an ISO date (YYYY-MM-DD); left unparsed",
        )
    try:
        _dt.date.fromisoformat(cleaned)
    except ValueError:
        return _CellResult(
            value=None,
            warning=f"{spec.name}: {text!r} is not a valid calendar date; left unparsed",
        )
    return _CellResult(value=cleaned)


def _coerce_cell(text: str, spec: ColumnSpec) -> _CellResult:
    if spec.kind == "number":
        return _coerce_number(text, spec)
    if spec.kind == "date":
        return _coerce_date(text, spec)
    return _CellResult(value=text.strip())


# ---------------------------------------------------------------------------
# Template-driven parse
# ---------------------------------------------------------------------------


def _build_header_index(columns: Sequence[ColumnSpec]) -> dict[str, ColumnSpec]:
    index: dict[str, ColumnSpec] = {}
    for spec in columns:
        for token in (spec.name, spec.header, *spec.aliases):
            index[normalize_header(token)] = spec
    return index


def parse_table(
    data: bytes,
    filename: str,
    columns: Sequence[ColumnSpec],
    *,
    kind: str,
    provenance: str,
    encoding: str | None = None,
) -> ParseReport:
    """Parse an uploaded file against a column contract into a :class:`ParseReport`.

    Behaviour (all non-fatal problems are collected, never raised):

    * A **missing required column** produces a clear error naming the column and no
      records (the sheet cannot be interpreted without it).
    * An **unknown/extra column** is ignored with a warning.
    * A **blank required cell** rejects only that row.
    * An **out-of-range** numeric value rejects that row with a message naming the
      specific allowed range; it is never silently imported.
    * An **ambiguous / unit-bearing** numeric value warns and is left unparsed.
    * A **single bad cell never discards the other rows**.

    Args:
        data: Raw uploaded bytes.
        filename: Original filename (selects CSV vs XLSX; ``.xlsm`` is rejected).
        columns: The template's column contract.
        kind: Domain label recorded on the report (e.g. ``"equipment"``).
        provenance: Provenance stamped on every emitted record.
        encoding: Optional explicit CSV encoding.

    Returns:
        A :class:`ParseReport`. Fatal file problems (unreadable / ``.xlsm``) raise
        :class:`IngestError`.
    """
    report = ParseReport(
        kind=kind,
        provenance=provenance,
        columns=[c.to_contract() for c in columns],
    )

    table = read_rows(data, filename, encoding=encoding)
    for warn in table.warnings:
        report.add_warning(row=1, message=warn)

    if not table.headers:
        report.add_error(row=1, message="file is empty: no header row found")
        return report

    header_index = _build_header_index(columns)
    # Map each present column position -> ColumnSpec (or None for unknown columns).
    position_specs: list[ColumnSpec | None] = []
    seen: set[str] = set()
    for raw_header in table.headers:
        spec = header_index.get(normalize_header(raw_header))
        position_specs.append(spec)
        if spec is None:
            if raw_header.strip():
                report.add_warning(
                    row=1,
                    column=raw_header,
                    message=f"unknown column {raw_header!r} ignored (not in the {kind} template)",
                )
        else:
            seen.add(spec.name)

    missing = [c for c in columns if c.required and c.name not in seen]
    if missing:
        for spec in missing:
            report.add_error(
                row=1,
                column=spec.name,
                message=f"required column {spec.name!r} is missing from the uploaded {kind} file",
            )
        return report

    for offset, raw_row in enumerate(table.rows):
        # 1-based spreadsheet row number: header is row 1, first data row is row 2.
        row_no = offset + 2
        if all(cell.strip() == "" for cell in raw_row):
            continue  # skip fully-blank spacer rows silently
        _parse_row(report, columns, position_specs, raw_row, row_no)

    return report


def _parse_row(
    report: ParseReport,
    columns: Sequence[ColumnSpec],
    position_specs: Sequence[ColumnSpec | None],
    raw_row: Sequence[str],
    row_no: int,
) -> None:
    values: dict[str, Any] = {}
    raw_by_name: dict[str, str] = {}
    for i, spec in enumerate(position_specs):
        if spec is None:
            continue
        raw_by_name[spec.name] = raw_row[i] if i < len(raw_row) else ""

    row_has_error = False
    for spec in columns:
        raw_value = raw_by_name.get(spec.name, "")
        if raw_value.strip() == "":
            if spec.required:
                report.add_error(
                    row=row_no,
                    column=spec.name,
                    message=f"required value {spec.name!r} is blank; row {row_no} skipped",
                )
                row_has_error = True
            else:
                values[spec.name] = spec.default
            continue

        result = _coerce_cell(raw_value, spec)
        if result.error is not None:
            report.add_error(row=row_no, column=spec.name, message=result.error)
            row_has_error = True
            continue
        if result.warning is not None:
            report.add_warning(row=row_no, column=spec.name, message=result.warning)
            if spec.required:
                report.add_error(
                    row=row_no,
                    column=spec.name,
                    message=(
                        f"required value {spec.name!r} could not be parsed; "
                        f"row {row_no} skipped"
                    ),
                )
                row_has_error = True
            else:
                values[spec.name] = spec.default
            continue
        values[spec.name] = result.value

    if row_has_error:
        return
    values["provenance"] = report.provenance
    values["_row"] = row_no
    report.records.append(values)


# ---------------------------------------------------------------------------
# Template generation (from the same contract the parser enforces)
# ---------------------------------------------------------------------------


def render_template_csv(columns: Sequence[ColumnSpec], *, include_example: bool = True) -> str:
    """Render a downloadable CSV template directly from a column contract.

    Generating the template from the contract guarantees it can never drift from
    the parser. The optional sample row uses each column's documented ``example``
    value, so a freshly downloaded template round-trips cleanly through its own
    parser. Cells are passed through :func:`escape_formula` for defence in depth.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([escape_formula(c.header) for c in columns])
    if include_example:
        writer.writerow([escape_formula(c.example) for c in columns])
    return buffer.getvalue()
